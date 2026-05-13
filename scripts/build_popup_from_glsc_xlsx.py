"""
Regenerate SLEW-popup-parameters.json from GLSC_CwaData.xlsx.
Usage: python scripts/build_popup_from_glsc_xlsx.py [path-to-GLSC_CwaData.xlsx]
"""
from __future__ import annotations

import json
import sys
from pathlib import Path

import openpyxl

REPO_ROOT = Path(__file__).resolve().parents[1]
OUT_JSON = REPO_ROOT / "SLEW-popup-parameters.json"
DEFAULT_XLSX = Path(r"e:\Users\Alex\Downloads\GLSC_CwaData.xlsx")

# Map XLSX ParameterID -> Clewa API variableTerm / variable aliases (case-insensitive match in HTML).
API_ALIASES: dict[str, list[str]] = {
    "encrh": ["Encrh", "encRH", "encrh"],
    "enctemp": ["Enctemp", "encTemp", "enctemp"],
    "airtemp": ["Air temperature", "airtemp", "airTemp"],
    "dewpt": ["Dew point", "dewpt"],
    "rh": ["Relative humidity", "rh"],
    "heatindex": ["Heat index", "heatIndex", "heatindex"],
    "baro": ["Barometric pressure", "baro"],
    "sbar": ["Sbar", "sbar"],
    "wspd": ["Wind speed", "wspd"],
    "wgst": ["Wind gust", "wgst"],
    "wdir": ["Wind direction", "wdir"],
    "wdirraw": ["Wdirraw", "wdirraw", "Wdirraw2"],
    "precipaccum": ["Precipitation accumulation", "precipAccum", "precipaccum"],
    "precipint": ["Precipint", "precipInt", "precipint"],
    "srad": ["Solar radiation", "srad"],
    "par": ["Par", "par"],
    "radmv": ["Radmv", "radmV", "radmv"],
    "soiltemp": ["Soiltemp 0.035m", "Soiltemp 0.1m", "Soiltemp 0.2m", "soilTemp", "soiltemp"],
    "soilea": ["SoilEa", "soilEa", "soilea"],
    "soilvwc": ["Soilvwc 0.075m", "Soilvwc 0.225m", "Soilvwc 0.375m", "soilVWC", "soilvwc"],
    "wvhgt": ["Wave height", "wvhgt"],
    "h10wvhgt": ["H10wvhgt", "h10wvhgt"],
    "maxwvhgt": ["Maxwvhgt", "maxWvhgt", "maxwvhgt"],
    "dompd": ["Dominant period", "dompd"],
    "maxpd": ["Maxpd", "maxpd"],
    "spcond": ["Spcond", "spCond", "spcond", "Spcond 6", "Spcond 8"],
    "turb": ["Turb", "turb", "Turb 6", "Turb 8"],
    "ph": ["Ph", "pH", "Ph 6", "Ph 8", "Ph 15"],
    "orp": ["Orp", "ORP", "Orp 6", "Orp 8"],
    "chlrfu": ["Chlrfu", "chlrfu", "Chlrfu 13.8", "Chlrfu 15"],
    "bgarfu": ["Bgarfu", "bgarfu", "Bgacells"],
    "do": ["Do", "do", "Do 6", "Do 8", "Do 15", "Do 21m"],
    "dosat": ["Dosat", "dosat", "Dosat 6", "Dosat 8"],
    "wligld": ["Water Level (ft) at IGLD85 - Sensor #1", "Water level", "wligld"],
    "wtemp": [
        "Water temperature",
        "Water temperature (0 m)",
        "wtemp",
        "wTemp",
        "Wtemp 1",
        "Wtemp 1m",
        "Wtemp 2",
        "Wtemp 2m",
        "Wtemp 6",
        "Wtemp 8",
        "Wtemp 10m",
        "Wtemp 12m",
        "Wtemp 15",
        "Wtemp 21m",
    ],
    "wtemp_0m": [
        "Water temperature",
        "wtemp",
        "Water temperature (0 m)",
        "Wtemp 1m",
        "Wtemp 1",
        "Wtemp 2m",
    ],
    "soilvwc-075": ["Soilvwc 0.075m"],
    "soilvwc-225": ["Soilvwc 0.225m"],
    "soilvwc-375": ["Soilvwc 0.375m"],
    "soiltemp-035": ["Soiltemp 0.035m"],
    "soiltemp-1": ["Soiltemp 0.1m"],
    "soiltemp-2": ["Soiltemp 0.2m"],
}


def pid_to_id(pid: str) -> str:
    p = str(pid).strip()
    fixed = {
        "encRH": "encrh",
        "encTemp": "enctemp",
        "airTemp": "airtemp",
        "heatIndex": "heatindex",
        "precipAccum": "precipaccum",
        "precipInt": "precipint",
        "radmV": "radmv",
        "soilTemp": "soiltemp",
        "soilEa": "soilea",
        "soilVWC": "soilvwc",
        "h10wvhgt": "h10wvhgt",
        "maxWvhgt": "maxwvhgt",
        "maxpd": "maxpd",
        "spCond": "spcond",
        "pH": "ph",
        "ORP": "orp",
        "do": "do",
        "dosat": "dosat",
        "chlrfu": "chlrfu",
        "bgarfu": "bgarfu",
        "wligld": "wligld",
    }
    if p in fixed:
        return fixed[p]
    return p.lower()


def fmt_row(def_id: str) -> tuple[str, dict]:
    d = def_id.lower()
    x: dict = {}
    if d in ("encrh", "rh", "dosat"):
        x["decimals"] = 1
        return "percent", x
    if d in ("enctemp", "airtemp", "dewpt", "wtemp", "wtemp_0m") or d.startswith("soiltemp"):
        x["hideIfLe"] = -99
        return "celsiusF", x
    if d.startswith("soilvwc"):
        x["decimals"] = 1
        return "number", x
    if d in ("wspd", "wgst"):
        x["hideIfLe"] = -99
        return "knotsFromMps", x
    if d in ("wdir", "wdirraw"):
        x["hideIfLe"] = -99
        return "degrees", x
    if d in ("wvhgt", "h10wvhgt", "maxwvhgt"):
        x["decimals"] = 2
        return "meters", x
    if d in ("dompd", "maxpd"):
        x["decimals"] = 0
        return "seconds", x
    if d == "srad":
        return "integer", x
    x["decimals"] = 2
    return "number", x


def terms_for_id(def_id: str, sheet_pid: str | None) -> list[str]:
    if def_id in API_ALIASES:
        raw = list(API_ALIASES[def_id])
    else:
        raw = [def_id, sheet_pid] if sheet_pid else [def_id]
    out, seen = [], set()
    for t in raw:
        if t is None:
            continue
        s = str(t).strip()
        if s and s.lower() not in seen:
            seen.add(s.lower())
            out.append(s)
    return out


def mou_token_to_ids(token: str) -> list[str]:
    t = token.strip().lower().replace(" ", "")
    if t in ("soilwvc", "soilvwc"):
        return ["soilvwc-075", "soilvwc-225", "soilvwc-375"]
    if t == "soiltemp":
        return ["soiltemp-035", "soiltemp-1", "soiltemp-2"]
    if t in ("aritemp", "airtemp"):
        return ["airtemp"]
    if t == "wtemp_0m":
        return ["wtemp_0m"]
    if t == "wtemp":
        return ["wtemp"]
    if t == "dompd":
        return ["dompd"]
    if t == "wgst":
        return ["wgst"]
    # normalize camel from MOU to id
    for pid, sid in [
        ("encrh", "encrh"),
        ("enctemp", "enctemp"),
        ("dewpt", "dewpt"),
        ("rh", "rh"),
        ("precipaccum", "precipaccum"),
        ("precipint", "precipint"),
        ("srad", "srad"),
        ("baro", "baro"),
        ("wspd", "wspd"),
        ("wdir", "wdir"),
        ("wvhgt", "wvhgt"),
        ("wligld", "wligld"),
    ]:
        if t == pid:
            return [sid]
    # fallback: try as ParameterID-like
    return [pid_to_id(token)]


def main() -> None:
    xlsx = Path(sys.argv[1]) if len(sys.argv) > 1 else DEFAULT_XLSX
    if not xlsx.is_file():
        print("XLSX not found:", xlsx)
        sys.exit(1)

    wb = openpyxl.load_workbook(xlsx, read_only=True, data_only=True)
    ws_p = wb["Parameter Names"]
    sheet_rows: list[tuple[str, str]] = []
    for row in ws_p.iter_rows(min_row=2, values_only=True):
        pid, desc = row[0], row[1]
        if not pid:
            break
        sheet_rows.append((str(pid).strip(), str(desc).strip() if desc else str(pid)))

    definitions: list[dict] = []
    for pid, desc in sheet_rows:
        sid = pid_to_id(pid)
        fmt, extra = fmt_row(sid)
        definitions.append(
            {
                "id": sid,
                "terms": terms_for_id(sid, pid),
                "label": desc,
                "format": fmt,
                **extra,
            }
        )

    # MOU-only parameters not on Parameter Names sheet; drop generic soil rows (depth ids below).
    definitions = [d for d in definitions if d["id"] not in ("soilvwc", "soiltemp")]

    extra_defs = [
        ("wgst", "Wind gust"),
        ("wtemp", "Water temperature"),
        ("wtemp_0m", "Water temperature (0 m)"),
        ("soilvwc-075", "Soil VWC (0.075 m)"),
        ("soilvwc-225", "Soil VWC (0.225 m)"),
        ("soilvwc-375", "Soil VWC (0.375 m)"),
        ("soiltemp-035", "Soil temperature (0.035 m)"),
        ("soiltemp-1", "Soil temperature (0.1 m)"),
        ("soiltemp-2", "Soil temperature (0.2 m)"),
    ]
    have = {d["id"] for d in definitions}
    for sid, lbl in extra_defs:
        if sid not in have:
            have.add(sid)
            fmt, extra = fmt_row(sid)
            definitions.append(
                {
                    "id": sid,
                    "terms": terms_for_id(sid, None),
                    "label": lbl,
                    "format": fmt,
                    **extra,
                }
            )

    ws_s = wb["Stations"]
    station_parameter_order: dict[str, list[str]] = {}
    for row in ws_s.iter_rows(min_row=2, values_only=True):
        name, label, mou = row[2], row[3], row[7]
        if mou is None or str(mou).strip() == "":
            continue
        ids: list[str] = []
        for part in str(mou).split(","):
            part = part.strip()
            if not part:
                continue
            for did in mou_token_to_ids(part):
                if did not in ids:
                    ids.append(did)
        keys: list[str] = []
        if name is not None and str(name).strip():
            keys.append(str(name).strip())
        if label is not None:
            keys.append(str(label).strip())
            if isinstance(label, (int, float)) and not isinstance(label, bool):
                keys.append(str(int(label)))
        for k in keys:
            if k and k not in station_parameter_order:
                station_parameter_order[k] = list(ids)

    wb.close()

    # API / fallback coords use lowercase device name for this site.
    if "Mwchargrin" in station_parameter_order and "mwchagrin" not in station_parameter_order:
        station_parameter_order["mwchagrin"] = list(station_parameter_order["Mwchargrin"])

    out = {
        "_comment": "Generated from GLSC_CwaData.xlsx (Sheets: Stations, Parameter Names). Rebuild: python scripts/build_popup_from_glsc_xlsx.py",
        "parameterDefinitions": definitions,
        "stationParameterOrder": station_parameter_order,
    }
    OUT_JSON.write_text(json.dumps(out, indent=2) + "\n", encoding="utf-8")
    print("Wrote", OUT_JSON)
    print("definitions:", len(definitions), "station keys:", len(station_parameter_order))


if __name__ == "__main__":
    main()
